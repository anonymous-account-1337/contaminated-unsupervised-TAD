from openpyxl.utils import get_column_letter


def adjust_worksheet_column_width(worksheet, df, index=True):
    if index:
        max_len = df.index.astype(str).map(len).max()
        worksheet.column_dimensions[get_column_letter(1)].width = max_len + 2

    for i, col in enumerate(df.columns, start=2 if index else 1):
        max_len = max(df[col].astype(str).map(len).max(), len(str(col)))
        worksheet.column_dimensions[get_column_letter(i)].width = max_len + 2
